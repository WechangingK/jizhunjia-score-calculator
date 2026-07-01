# -*- coding: utf-8 -*-
"""Excel导入导出工具

支持:
- 价格计算：导入应答人报价、导出价格得分结果
- 综合评审：导入供应商综合数据、导出评审结果
"""

from typing import Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.calculator import CalcResult


class ExcelIO:
	"""Excel导入导出"""

	NAME_PATTERNS = [
		'应答人名称', '公司名称', '供应商名称', '应答人', '名称',
		'公司', '供应商', '投标人', '投标人名称', '单位名称', '单位'
	]

	PRICE_PATTERNS = [
		'不含税总价', '报价', '总价', '价格', '不含税价',
		'投标报价', '应答报价', '金额', '总报价', '不含税'
	]

	@staticmethod
	def importBidders(filePath: str) -> Dict[str, float]:
		"""从Excel文件导入应答人数据"""
		df = pd.read_excel(filePath)

		nameCol = None
		for col in df.columns:
			colStr = str(col).strip()
			for pattern in ExcelIO.NAME_PATTERNS:
				if pattern in colStr:
					nameCol = col
					break
			if nameCol is not None:
				break

		if nameCol is None:
			raise ValueError(
				f'无法识别名称列，请确保Excel包含以下列名之一：\n'
				f'{", ".join(ExcelIO.NAME_PATTERNS[:5])}'
			)

		priceCol = None
		for col in df.columns:
			colStr = str(col).strip()
			for pattern in ExcelIO.PRICE_PATTERNS:
				if pattern in colStr:
					priceCol = col
					break
			if priceCol is not None:
				break

		if priceCol is None:
			raise ValueError(
				f'无法识别报价列，请确保Excel包含以下列名之一：\n'
				f'{", ".join(ExcelIO.PRICE_PATTERNS[:5])}'
			)

		bidders = {}
		for _, row in df.iterrows():
			name = str(row[nameCol]).strip()
			price = row[priceCol]
			if not name or name.lower() == 'nan':
				continue
			try:
				price = float(price)
			except (ValueError, TypeError):
				continue
			bidders[name] = price

		if not bidders:
			raise ValueError('未读取到有效数据，请检查Excel内容')

		return bidders

	@staticmethod
	def exportResults(
		calcResult: CalcResult,
		ruleName: str,
		filePath: str
	):
		"""导出计算结果到Excel文件

		Args:
			calcResult: 完整计算结果
			ruleName: 使用的规则名称
			filePath: 导出路径
		"""
		wb = Workbook()
		ws = wb.active
		ws.title = '价格得分计算结果'

		# 样式定义
		headerFill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
		headerFont = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
		dataFont = Font(name='微软雅黑', size=10)
		redFont = Font(name='微软雅黑', size=10, color='FF0000')
		greenFont = Font(name='微软雅黑', size=10, color='008000', bold=True)
		centerAlign = Alignment(horizontal='center', vertical='center')
		thinBorder = Border(
			left=Side(style='thin'), right=Side(style='thin'),
			top=Side(style='thin'), bottom=Side(style='thin')
		)

		rule = calcResult.rule
		limitText = f'{rule.maxPrice:,.2f} 元' if rule and rule.maxPrice > 0 else '不限价'
		fullScore = rule.fullScore if rule else 80

		# 标题行
		ws.merge_cells('A1:H1')
		titleCell = ws['A1']
		titleCell.value = '基准价得分计算结果'
		titleCell.font = Font(name='微软雅黑', bold=True, size=14)
		titleCell.alignment = Alignment(horizontal='center', vertical='center')

		ws.merge_cells('A2:H2')
		infoCell = ws['A2']
		infoCell.value = (
			f'评分规则：{ruleName}　|　'
			f'基准价：{calcResult.benchmark:,.2f} 元　|　'
			f'满分：{fullScore} 分　|　'
			f'最高限价：{limitText}　|　'
			f'有效应答人数：{calcResult.validCount}'
			+ (f'　|　排除：{calcResult.excludedCount}' if calcResult.excludedCount > 0 else '')
		)
		infoCell.font = Font(name='微软雅黑', size=10)
		infoCell.alignment = Alignment(horizontal='center', vertical='center')

		# 表头
		headers = ['排名', '应答人名称', '不含税总价', '基准价', '偏离基准价(%)',
				   f'价格得分(满分{fullScore})', '有效性', '备注']
		for colIdx, header in enumerate(headers, 1):
			cell = ws.cell(row=4, column=colIdx, value=header)
			cell.font = headerFont
			cell.fill = headerFill
			cell.alignment = centerAlign
			cell.border = thinBorder

		# 数据行
		for rowIdx, r in enumerate(calcResult.results):
			rowNum = 5 + rowIdx

			validText = '有效' if r.valid else '无效'

			if r.valid:
				rankText = str(r.rank)
				devText = r.deviation
				scoreText = r.score
				if r.deviation > 0:
					note = '高于基准价'
				elif r.deviation < 0:
					note = '低于基准价'
				else:
					note = '等于基准价'
			else:
				rankText = '—'
				devText = '—'
				scoreText = 0.0
				note = r.invalidReason

			data = [rankText, r.name, r.price, calcResult.benchmark,
					devText, scoreText, validText, note]

			isInvalid = not r.valid
			isFullScore = r.valid and r.score >= fullScore
			isZeroScore = r.valid and r.score <= 0

			for colIdx, value in enumerate(data, 1):
				cell = ws.cell(row=rowNum, column=colIdx, value=value)
				cell.font = dataFont
				cell.alignment = centerAlign
				cell.border = thinBorder

				if isInvalid:
					cell.font = redFont
				elif colIdx == 6:  # 得分列
					if isFullScore:
						cell.font = greenFont
					elif isZeroScore:
						cell.font = redFont

		# 列宽
		colWidths = [6, 20, 14, 14, 16, 16, 8, 26]
		for colIdx, width in enumerate(colWidths, 1):
			colLetter = chr(64 + colIdx)
			ws.column_dimensions[colLetter].width = width

		wb.save(filePath)

	# ==================== 综合评审导入导出 ====================

	EVAL_NAME_PATTERNS = [
		'供应商名称', '公司名称', '应答人名称', '供应商', '名称',
		'公司', '单位名称', '单位',
	]

	EVAL_COL_MAP = {
		'业绩': 'contractCount', '合同': 'contractCount', '类似项目': 'contractCount',
		'团队': 'teamCount', '人员': 'teamCount',
		'故障': 'faultHours', '时限': 'faultHours', '处理时间': 'faultHours',
		'整体服务': 'servicePlanScore', '服务方案': 'servicePlanScore',
		'质量': 'qualityPlanScore', '质量控制': 'qualityPlanScore',
		'应急': 'emergencyPlanScore', '应急响应': 'emergencyPlanScore',
		'限制': 'isRestricted', '扣分': 'isRestricted',
		'报价': 'price', '价格': 'price', '总价': 'price', '不含税': 'price',
	}

	@staticmethod
	def importEvaluationData(filePath: str) -> List:
		"""从Excel导入综合评审供应商数据

		自动识别列名映射到 SupplierInput 字段。

		Args:
			filePath: Excel文件路径

		Returns:
			SupplierInput 列表
		"""
		from core.supplier_scorer import SupplierInput

		df = pd.read_excel(filePath)

		# 自动映射列
		colMapping = {}
		for col in df.columns:
			colStr = str(col).strip()
			# 名称列
			for pat in ExcelIO.EVAL_NAME_PATTERNS:
				if pat in colStr:
					colMapping['name'] = col
					break
			# 其他列
			if 'name' in colMapping and colMapping['name'] == col:
				continue
			for keyword, field in ExcelIO.EVAL_COL_MAP.items():
				if keyword in colStr:
					colMapping[field] = col
					break

		if 'name' not in colMapping:
			raise ValueError(
				f'无法识别供应商名称列，请确保Excel包含以下列名之一：\n'
				f'{", ".join(ExcelIO.EVAL_NAME_PATTERNS[:5])}'
			)

		suppliers = []
		for _, row in df.iterrows():
			name = str(row[colMapping['name']]).strip()
			if not name or name.lower() == 'nan':
				continue

			supplier = SupplierInput(name=name)

			for field, col in colMapping.items():
				if field == 'name':
					continue
				val = row[col]
				try:
					if field == 'isRestricted':
						setattr(supplier, field, bool(val))
					elif field in ('contractCount', 'teamCount'):
						setattr(supplier, field, int(float(val)))
					elif field == 'price':
						setattr(supplier, field, float(val) if val and str(val).lower() != 'nan' else 0.0)
					else:
						setattr(supplier, field, float(val) if val and str(val).lower() != 'nan' else 0.0)
				except (ValueError, TypeError):
					pass

			suppliers.append(supplier)

		return suppliers

	@staticmethod
	def exportEvaluationResults(report, ruleName: str, filePath: str):
		"""导出综合评审结果到Excel

		Args:
			report: ComparisonReport 对象
			ruleName: 价格规则名称
			filePath: 导出路径
		"""
		from core.comparison_engine import ComparisonReport

		wb = Workbook()

		# ---- Sheet1: 综合评分结果 ----
		ws = wb.active
		ws.title = '综合评分结果'

		# 样式
		headerFill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
		headerFont = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
		dataFont = Font(name='微软雅黑', size=10)
		boldFont = Font(name='微软雅黑', size=10, bold=True)
		greenFont = Font(name='微软雅黑', size=10, color='008000', bold=True)
		redFont = Font(name='微软雅黑', size=10, color='FF0000')
		centerAlign = Alignment(horizontal='center', vertical='center')
		thinBorder = Border(
			left=Side(style='thin'), right=Side(style='thin'),
			top=Side(style='thin'), bottom=Side(style='thin')
		)

		# 标题行
		ws.merge_cells('A1:K1')
		titleCell = ws['A1']
		titleCell.value = '综合评审计算结果'
		titleCell.font = Font(name='微软雅黑', bold=True, size=14)
		titleCell.alignment = Alignment(horizontal='center', vertical='center')

		# 信息行
		ws.merge_cells('A2:K2')
		infoCell = ws['A2']
		winnerName = report.winner.name if report.winner else '—'
		infoCell.value = (
			f'价格规则：{ruleName}　|　'
			f'参评供应商：{len(report.results)}家　|　'
			f'综合最优：{winnerName}'
		)
		infoCell.font = Font(name='微软雅黑', size=10)
		infoCell.alignment = Alignment(horizontal='center', vertical='center')

		# 表头
		headers = ['排名', '供应商名称', '商务/业绩(8)', '团队人员(6)',
				   '故障时限(2)', '整体服务(3)', '质量控制(3)',
				   '应急响应(3)', '扣分', '价格(75)', '总分(100)']
		for colIdx, header in enumerate(headers, 1):
			cell = ws.cell(row=4, column=colIdx, value=header)
			cell.font = headerFont
			cell.fill = headerFill
			cell.alignment = centerAlign
			cell.border = thinBorder

		# 数据行
		for rowIdx, r in enumerate(report.results):
			rowNum = 5 + rowIdx

			isWinner = (r.rank == 1)
			fontToUse = boldFont if isWinner else dataFont

			fs = r.factorScores
			data = [
				r.rank, r.name,
				fs.get(1, 0),
				fs.get(2, 0),
				fs.get(3, 0),
				fs.get(4, 0),
				fs.get(5, 0),
				fs.get(6, 0),
				fs.get(7, 0),
				fs.get(8, 0),
				r.totalScore,
			]

			for colIdx, value in enumerate(data, 1):
				cell = ws.cell(row=rowNum, column=colIdx, value=value)
				cell.font = fontToUse
				cell.alignment = centerAlign
				cell.border = thinBorder

				if colIdx == 9 and value < 0:
					cell.font = redFont
				if colIdx == 11:
					cell.font = greenFont if isWinner else fontToUse

		# 列宽
		colWidths = [6, 20, 12, 10, 10, 10, 10, 10, 8, 10, 12]
		for colIdx, width in enumerate(colWidths, 1):
			ws.column_dimensions[get_column_letter(colIdx)].width = width

		# ---- Sheet2: 智能分析建议 ----
		ws2 = wb.create_sheet('智能分析建议')

		ws2.merge_cells('A1:D1')
		ws2['A1'].value = '智能分析与建议'
		ws2['A1'].font = Font(name='微软雅黑', bold=True, size=14)
		ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')

		row = 3
		ws2.merge_cells(f'A{row}:D{row}')
		ws2.cell(row=row, column=1, value='📊 总体概况').font = Font(name='微软雅黑', bold=True, size=11)
		row += 1
		ws2.merge_cells(f'A{row}:D{row}')
		ws2.cell(row=row, column=1, value=report.summary).font = Font(name='微软雅黑', size=10)
		ws2.row_dimensions[row].height = 40

		if report.recommendations:
			row += 2
			ws2.merge_cells(f'A{row}:D{row}')
			ws2.cell(row=row, column=1, value='🎯 综合建议').font = Font(name='微软雅黑', bold=True, size=11)
			for rec in report.recommendations:
				row += 1
				ws2.merge_cells(f'A{row}:D{row}')
				ws2.cell(row=row, column=1, value=f'• {rec}').font = Font(name='微软雅黑', size=10)

		row += 2
		ws2.merge_cells(f'A{row}:D{row}')
		ws2.cell(row=row, column=1, value='📋 逐家分析').font = Font(name='微软雅黑', bold=True, size=11)
		for name, analysis in report.analysisBySupplier.items():
			row += 1
			ws2.merge_cells(f'A{row}:D{row}')
			ws2.cell(row=row, column=1, value=analysis.replace('\n', ' | ')).font = Font(name='微软雅黑', size=10)
			ws2.row_dimensions[row].height = 30

		for colIdx in range(1, 5):
			ws2.column_dimensions[get_column_letter(colIdx)].width = 30

		wb.save(filePath)
